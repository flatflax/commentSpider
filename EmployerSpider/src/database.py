# coding=utf-8

import sqlite3, logging, os, queue
from openpyxl import Workbook, load_workbook

from src import settings, tools


class Database():
    def __init__(self):
        self.database_path = settings.database_path

    def get_connect(self):
        conn = sqlite3.connect(self.database_path)
        return conn

    def create_table(self, table_name):
        connect = sqlite3.connect(self.database_path)
        cur = connect.cursor()
        cur.execute("select count(*) from sqlite_master where type='table' and name='{}'".format(table_name))
        value = cur.fetchall()
        if value[0][0] != 1:
            create_sql = self.get_create_sql(table_name)
            cur.execute(create_sql)
            logging.info("table {} create".format(table_name))
        else:
            logging.info("table {} exist".format(table_name))
        connect.commit()
        connect.close()

    def export_database_data(self, table_name):
        select_sql, name_list = self.get_select_sql(table_name)
        conn = sqlite3.connect(self.database_path)
        cursor = conn.cursor()
        cursor.execute(select_sql)
        names = list(map(lambda x: x[0], cursor.description))
        result = cursor.fetchall()

        file_path = settings.export_company_file.format(date=tools.get_today(), table_name = table_name)
        if len(result) > 0:
            if not os.path.exists(file_path):
                wb = Workbook()
                ws = wb.active
                ws.append(name_list)
                ws.append(names)
            else:
                wb = load_workbook(file_path)
                ws = wb.get_active_sheet()
                if not os.path.getsize(file_path):
                    ws.append(names)
            for r in result:
                ws.append(r)
            wb.save(file_path)

    def get_select_sql(self, table_name):
        sql, name_list = "", tuple([])
        if table_name == 'job_info':
            sql = "SELECT positionId, positionTitle, positionDistrict, minPositionSalary, " \
                  "maxPositionSalary, bonusPositionSalary, companyName,companyId,companyDistrict," \
                  "companyAddress,companyLocation, website FROM job_info"
            name_list = tuple(['职位id', '职位名称', '职位地区', '薪资最小值', '薪资最大值', '一年x薪',
                               '公司名称', '公司Id', '公司地区', '地址', '经纬度', '平台'])
        if table_name == 'company_info':
            sql = "SELECT companyId, companyName, companyDistrict, " \
                  "companyAddress, companyLocation, companyHireNum, website " \
                  "FROM company_info"
            name_list = tuple(['公司id','公司名','公司地区','公司地址','公司经纬度','公司招聘职位数','平台'])
        return sql, name_list

    def get_create_sql(self, table_name):
        sql = ''
        if table_name == 'job_info':
            '''
             positionId  职位id
             positionTitle 职位名称
             positionDistrict 职位地区
             minPositionSalary 薪资最小值
             maxPositionSalary 薪资最大值
             bonusPositionSalary 一年x薪
             companyName 公司名
             companyId 公司id(招聘网站页)
             companyDistrict 公司地区
             companyAddress 公司地址
             companyLocation 公司经纬度
             website 平台
            '''
            sql = "create table job_info (" \
                  "positionId varchar(50) PRIMARY KEY," \
                  "positionTitle varchar(50)," \
                  "positionDistrict varchar(50)," \
                  "minPositionSalary varchar(5)," \
                  "maxPositionSalary varchar(5)," \
                  "bonusPositionSalary varchar(3)," \
                  "companyName varchar(30), " \
                  "companyId varchar(30), " \
                  "companyDistrict varchar(30), " \
                  "companyAddress varchar(100)," \
                  "companyLocation varchar(30)," \
                  "website varchar(10)" \
                  ")"
        if table_name == 'company_info':
            '''
            companyId  公司id
            companyName 公司名
            companyDistrict 公司地区
            companyAddress 公司地址
            companyLocation 公司经纬度
            companyHireNum 公司招聘职位数
            website 平台
            '''
            sql = "create table company_info (" \
                  "companyId varchar(30) PRIMARY KEY," \
                  "companyName varchar(30), " \
                  "companyDistrict varchar(30), " \
                  "companyAddress varchar(100)," \
                  "companyLocation varchar(30)," \
                  "companyHireNum int(5)," + \
                  "website varchar(10)" \
                  ")"
        return sql


class PoolException(Exception):
    pass


class Pool(object):
    """一个数据库连接池"""

    def __init__(self, maxActive=5, maxWait=None, init_size=0, db_type="SQLite3", **config):
        self.__freeConns = queue.Queue(maxActive)
        self.maxWait = maxWait
        self.db_type = db_type
        self.config = config
        if init_size > maxActive:
            init_size = maxActive
        for i in range(init_size):
            self.free(self._create_conn())

    def __del__(self):
        print("__del__ Pool..")
        self.release()

    def release(self):
        '''释放资源，关闭池中的所有连接'''
        print("release Pool..")
        while self.__freeConns and not self.__freeConns.empty():
            con = self.get()
            con.release()
        self.__freeConns = None

    def _create_conn(self):
        '''创建连接 '''
        if self.db_type in dbcs:
            return dbcs[self.db_type](**self.config);

    def get(self, timeout=None):
        '''获取一个连接
        @param timeout:超时时间
        '''
        if timeout is None:
            timeout = self.maxWait
        conn = None
        if self.__freeConns.empty():  # 如果容器是空的，直接创建一个连接
            conn = self._create_conn()
        else:
            conn = self.__freeConns.get(timeout=timeout)
        conn.pool = self
        return conn

    def free(self, conn):
        '''将一个连接放回池中
        @param conn: 连接对象
        '''
        conn.pool = None
        if (self.__freeConns.full()):  # 如果当前连接池已满，直接关闭连接
            conn.release()
            return
        self.__freeConns.put_nowait(conn)


from abc import ABCMeta, abstractmethod


class PoolingConnection(object):
    def __init__(self, **config):
        self.conn = None
        self.config = config
        self.pool = None

    def __del__(self):
        self.release()

    def __enter__(self):
        pass

    def __exit__(self, exc_type, exc_value, traceback):
        self.close()

    def release(self):
        print("release PoolingConnection..")
        if self.conn is not None:
            self.conn.close()
            self.conn = None
        self.pool = None

    def close(self):
        if self.pool is None:
            raise PoolException("连接已关闭")
        self.pool.free(self)

    def __getattr__(self, val):
        if self.conn is None and self.pool is not None:
            self.conn = self._create_conn(**self.config)
        if self.conn is None:
            raise PoolException("无法创建数据库连接 或连接已关闭")
        return getattr(self.conn, val)

    @abstractmethod
    def _create_conn(self, **config):
        pass


class SQLit3PoolConnection(PoolingConnection):
    def _create_conn(self, **config):
        import sqlite3
        return sqlite3.connect(**config)


dbcs = {"SQLite3": SQLit3PoolConnection}

# pool = Pool(database="employer_database.db")
#
#
# def test():
#     conn = pool.get()
#     with conn:
#         for a in conn.execute("SELECT * FROM A"):
#             print(a)
