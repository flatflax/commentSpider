# coding=utf-8

import sqlite3, logging, os, queue
from openpyxl import Workbook, load_workbook

from src import settings, tools


class Database():
    def create_company_job_table(self, connect):
        cur = connect.cursor()
        cur.execute("select count(*) from sqlite_master where type='table' and name='company_job'")
        value = cur.fetchall()
        if value[0][0] != 1:
            cur.execute("create table employer_job (" +
                        "_id varchar(30) PRIMARY KEY," +
                        "companyName varchar(30), " +  # 公司名
                        "companyId varchar(30), " +  # 公司id(招聘网站页)
                        "companyDistrict varchar(30), " +  # 公司地区
                        "companyAddress varchar(100)," +  # 公司地址
                        "companyLocation varchar(30)," +  # 公司经纬度
                        "website varchar(10)," +  # 平台
                        "positionTitle varchar(50)," + # 职位名称
                        "positionId varchar(50)," +  # 职位id
                        "positionDistrict varchar(50)" +  # 职位地区
                        ")")
            logging.info("table company_job create")
        else:
            logging.info("table company_job exist")
        connect.commit()
        connect.close()

    def export_database_customer_msg(self, conn):
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM company_job")
        names = list(map(lambda x: x[0], cursor.description))
        result = cursor.fetchall()

        file_path = settings.database_export_msg_xlsx.format(date=tools.get_today())
        if len(result) > 0:
            if not os.path.exists(file_path):
                wb = Workbook()
                ws = wb.active
                ws.append(names)
            else:
                wb = load_workbook(file_path)
                ws = wb.get_active_sheet()
                if not os.path.getsize(file_path):
                    ws.append(names)
            for r in result:
                ws.append(r)
            wb.save(file_path)


class PoolException(Exception):
    pass


class Pool(object):
    """一个数据库连接池"""
    def __init__(self, maxActive=5, maxWait=None, init_size=0, db_type="SQLite3", **config):
        self.__freeConns = queue.Queue(maxActive)
        self.maxWait = maxWait
        self.db_type = db_type
        self.config = config
        if init_size > maxActive:
            init_size = maxActive
        for i in range(init_size):
            self.free(self._create_conn())

    def __del__(self):
        print("__del__ Pool..")
        self.release()

    def release(self):
        '''释放资源，关闭池中的所有连接'''
        print("release Pool..")
        while self.__freeConns and not self.__freeConns.empty():
            con = self.get()
            con.release()
        self.__freeConns = None

    def _create_conn(self):
        '''创建连接 '''
        if self.db_type in dbcs:
            return dbcs[self.db_type](**self.config);

    def get(self, timeout=None):
        '''获取一个连接
        @param timeout:超时时间
        '''
        if timeout is None:
            timeout = self.maxWait
        conn = None
        if self.__freeConns.empty():  # 如果容器是空的，直接创建一个连接
            conn = self._create_conn()
        else:
            conn = self.__freeConns.get(timeout=timeout)
        conn.pool = self
        return conn

    def free(self, conn):
        '''将一个连接放回池中
        @param conn: 连接对象
        '''
        conn.pool = None
        if (self.__freeConns.full()):  # 如果当前连接池已满，直接关闭连接
            conn.release()
            return
        self.__freeConns.put_nowait(conn)


from abc import ABCMeta, abstractmethod


class PoolingConnection(object):
    def __init__(self, **config):
        self.conn = None
        self.config = config
        self.pool = None

    def __del__(self):
        self.release()

    def __enter__(self):
        pass

    def __exit__(self, exc_type, exc_value, traceback):
        self.close()

    def release(self):
        print("release PoolingConnection..")
        if self.conn is not None:
            self.conn.close()
            self.conn = None
        self.pool = None

    def close(self):
        if self.pool is None:
            raise PoolException("连接已关闭")
        self.pool.free(self)

    def __getattr__(self, val):
        if self.conn is None and self.pool is not None:
            self.conn = self._create_conn(**self.config)
        if self.conn is None:
            raise PoolException("无法创建数据库连接 或连接已关闭")
        return getattr(self.conn, val)

    @abstractmethod
    def _create_conn(self, **config):
        pass


class SQLit3PoolConnection(PoolingConnection):
    def _create_conn(self, **config):
        import sqlite3
        return sqlite3.connect(**config)


dbcs = {"SQLite3": SQLit3PoolConnection}

pool = Pool(database="employer_database.db")


def test():
    conn = pool.get()
    with conn:
        for a in conn.execute("SELECT * FROM A"):
            print(a)